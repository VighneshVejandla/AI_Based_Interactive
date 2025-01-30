# backend/excel_processor.py

import openpyxl

class ExcelProcessor:
    def __init__(self, file_path):
        """
        Initialize the Excel Processor with the path to the Excel file.
        """
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(file_path)
        self.sheet = self.workbook.active

    def recalculate_totals(self):
        """
        Recalculate totals, sub-totals, and grand totals in the Excel file.
        Add "Calculation" and "Difference" rows below each total.
        """
        try:
            # Iterate through the rows and columns to find totals
            for row in self.sheet.iter_rows():
                for cell in row:
                    if "Total" in str(cell.value):  # Check if the cell contains a total
                        # Get the range of cells to sum (assuming the total is at the end of a row)
                        start_col = 1  # Start from the first column
                        end_col = cell.column - 1  # End at the column before the total
                        sum_range = self.sheet.iter_cols(min_col=start_col, max_col=end_col, min_row=cell.row, max_row=cell.row)

                        # Calculate the sum of the range
                        total_sum = sum(float(cell[0].value) for cell in sum_range if cell[0].value is not None)

                        # Add a "Calculation" row below the total
                        self.sheet.cell(row=cell.row + 1, column=cell.column, value="Calculation")
                        self.sheet.cell(row=cell.row + 1, column=cell.column, value=total_sum)

                        # Add a "Difference" row below the calculation
                        difference = total_sum - float(cell.value)
                        self.sheet.cell(row=cell.row + 2, column=cell.column, value="Difference")
                        self.sheet.cell(row=cell.row + 2, column=cell.column, value=difference)

            # Save the updated workbook
            self.workbook.save(self.file_path)
            print(f"Recalculations completed and saved to {self.file_path}")
        except Exception as e:
            print(f"Error in Excel processing: {e}")

    def get_cell_value(self, row, col):
        """
        Get the value of a specific cell in the Excel sheet.
        
        :param row: The row number (1-based index).
        :param col: The column number (1-based index).
        :return: The value of the cell.
        """
        return self.sheet.cell(row=row, column=col).value